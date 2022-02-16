import pandas as pd
from openpyxl import Workbook
from airtest.core.api import *

# 读取excel
def read_excel(file, **kwargs):
    data_dict = []  # 列表
    try:
        data = pd.read_excel(file, **kwargs)  # 读取excel文件
        data_dict = data.to_dict('records')  # 转化为字典类型，但是输入“recorda”，转化后为列表形式
    finally:
        return data_dict  # 返回列表


# 写入excel操作
def write_excel(filename, sheet, data, head=None):
    if head is None:
        head = []
    wb = Workbook()

    # 写入表头
    excel_head = head
    sheet_name = sheet

    sheet0 = wb.create_sheet(sheet_name, index=0)
    for i, item in enumerate(excel_head):
        sheet0.cell(row=1, column=i + 1, value=item)

    # 写入数据
    for i, item in enumerate(data):
        i = i + 2
        for j, val in enumerate(item):
            sheet0.cell(row=i, column=j + 1, value=val)

    wb.save(filename + '.xlsx')
    log("写入数据成功！！！", desc="--excel--")


# from Comm.phone_data import ran_list
#
# filename_0 = r'D:/PycharmProjects/AirProjectDome/TestCase/Demo/TestData/test_0'
# head_0 = '序号', '手机号码'
# sheet_0 = 'data01'
# data_0 = ran_list(5)
#
# write_excel(filename_0, sheet_0, data_0, head_0)
